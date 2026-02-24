import socket
from threading import Thread
from concurrent.futures import ThreadPoolExecutor, as_completed, wait, FIRST_COMPLETED
import traceback
import os
import os.path
#import json
import orjson
import re
from datetime import datetime
import time
import getopt
import sys
from pathlib import Path
from rich.console import Console
from rich.logging import RichHandler
import logging
import platform
import netmiko
import openpyxl
import optparse

from time import perf_counter

from class_network_device import NetworkDevice

from functions_gatherers import (
    set_verbose,
    gather_version,
    gather_arp,
    gather_mac,
    gather_interface,
    gather_cdp,
    gather_lldp,
    gather_route,
    gather_bgp,
    gather_inventory,
    attach_wtp_to_lldp_and_update_remote_software,
    gather_wtp_status,
    apply_wtp_software_to_lldp,
    gather_commands,
    count_interfaces
)
from functions_initial_setup import (
    update_ntc_templ_path,
    get_json_data_from_file,
    get_setup_vars,
    read_network_devices,
    save_device_type_cache,
    load_device_type_cache,
)
from functions_helpers import (
    find_val_in_col,
    gen_spacer,
    map_headers,
    next_available_row,
    format_uptime,
    left,
    right,
    mid,
    center_string,
    rw_cell,
    add_xls_tag,
    get_xls_sheet,
    open_xls,
    save_xls,
    mod_dir_based_on_os,
    verify_path,
    print_net_dev_msg,
    get_current_time,
)

VERBOSE = False
VERBOSE_MORE = False
VERBOSITY_LEVEL = 0

def set_runtime(verbose=None, verbose_more=None, verbosity_level=None):
    global VERBOSE_MORE, VERBOSITY_LEVEL
    if verbose_more is not None:
        VERBOSE_MORE = verbose_more
    if verbosity_level is not None:
        VERBOSITY_LEVEL = verbosity_level

####Save Functions
def save_device_data(net_devices, wb_obj, setup_vars, key_map, headers_key, input_file_name, next_row_cache, all_devices=None):
    glbl_set = setup_vars["global"]
    settings = setup_vars["settings"]
    
    for net_dev in net_devices:
        start_time = perf_counter()

        net_dev.update_outdir_outfile(glbl_set["output_dir"])
        if net_dev.active == "Completed" and net_dev.elapsed_time:
            gather_results_to_wb(wb_obj, net_dev, key_map, headers_key, next_row_cache)
            if VERBOSE_MORE: #was VERBOSE
                delta_timer = perf_counter() - start_time
                print(f"{delta_timer:.3f}s gather_results_to_wb for: {net_dev.host}")
            if settings["gather_commands"]:
                start_time2 = perf_counter()
                save_other_shows_to_txt(net_dev)
                delta_timer2 = perf_counter() - start_time2
                if VERBOSE_MORE: #was VERBOSE
                    print(f"{delta_timer2:.3f}s save_other_shows_to_txt for: {net_dev.host}")

        delta_timer = perf_counter() - start_time
        if VERBOSE_MORE: #was VERBOSE
            print(f"{delta_timer:.3f}s gather_results_to_wb and save_other_shows_to_txt for: {net_dev.host}")

        write_dev_vars_to_wb(wb_obj, net_dev, key_map["device_info_map"])
        save_dev_show_json_data(net_dev)
        add_err_msgs_to_wb(net_dev, wb_obj)

        delta_timer = perf_counter() - start_time
        if VERBOSE_MORE: #was VERBOSE
            print(f"{delta_timer:.3f}s save_device_data for: {net_dev.host}")

    save_xls(wb_obj, input_file_name, glbl_set["output_file"], glbl_set["output_dir"], VERBOSITY_LEVEL)

def save_dev_show_json_data(net_dev):
    """
    Same 'pretty text' format as before (spacers + ****** cmd ****** + EOF),
    but writes bytes end-to-end so we can write orjson bytes directly
    and avoid decode() overhead.
    """
    start_time = perf_counter()
    json_data = net_dev.show_output_json
    if not json_data:
        return

    json_f_name = net_dev.json_out_file
    output_dir = verify_path(net_dev.out_dir + "JSON/")
    file_name = output_dir + json_f_name

    spacer_top = gen_spacer("-", 2)
    spacer_block = gen_spacer()
    cmd_output_spacer = gen_spacer("#", 1)

    # Helper: write a text string as UTF-8 with \n normalized
    def wline(fh, s: str) -> None:
        # normalize newlines to \n, then encode
        fh.write(s.replace("\r\n", "\n").replace("\r", "\n").encode("utf-8"))

    with open(file_name, "wb") as fh:
        # Header
        wline(fh, spacer_top + center_string("Connected to " + net_dev.host) + "\n")
        wline(fh, center_string("Hostname is: " + net_dev.hostname) + "\n")
        wline(fh, spacer_top)

        # Per-command blocks
        for show_cmd, output in json_data.items():
            wline(fh, spacer_block + center_string("****** " + show_cmd + " ******") + "\n")
            wline(fh, cmd_output_spacer)
            fh.write(orjson.dumps(output, option=orjson.OPT_INDENT_2))
            wline(fh, "\n" + cmd_output_spacer)

        # EOF
        wline(fh, spacer_block + "*" * 20 + "\tEnd of File\t" + "*" * 20)
    
    delta_timer = perf_counter() - start_time
    
    if VERBOSE_MORE: #was VERBOSE
        print(f"{delta_timer:.3f}s JSON write: {json_f_name}")

def write_dev_vars_to_wb(wb_obj, device, var_loc):
    """
    Writes all the data to the "Main" worksheet
    """
    wb_sheet = wb_obj["Main"]
    dev_vars = vars(device)
    row = device.main_col
    for key, col in var_loc.items():
        if key in list(dev_vars.keys()):
            if key == "interface_count" and dev_vars[key]:
                for interf, c_key in col.items():
                    values = dev_vars[key][interf]
                    rw_cell(wb_sheet, row, c_key["count"], True, str(values["count"]))
                    rw_cell(wb_sheet, row, c_key["active"], True, str(values["active"]))
            elif isinstance(col, int):
                rw_cell(wb_sheet, row, col, True, str(dev_vars[key]))

def save_other_shows_to_txt(net_dev):
    """
    Saves the results to the "other" show commands entered on the
    spreadsheet. creates a new file for every device in a subdirectory
    of the "output directory"
    """
    if net_dev.user_rqstd_show:
        file_name = verify_path(net_dev.out_dir + "show_cmd_captures/")
        spacer = gen_spacer()
        cmd_output_spacer = gen_spacer("#", 1)
        file_name += net_dev.hostname + ".txt"
        with open(file_name, "w+") as filehandle:
            write_str = spacer
            write_str += center_string("Connected to " + net_dev.host) + "\n"
            write_str += center_string("Hostname is: " + net_dev.hostname)
            write_str += "\n" + spacer
            filehandle.write(write_str)
            for show_cmd in net_dev.user_rqstd_show.keys():
                write_str = show_cmd + "\n"
                filehandle.write(write_str)
            filehandle.write(spacer)
            for show_cmd, output in net_dev.user_rqstd_show.items():
                write_str = center_string("****** " + show_cmd + " ******")
                write_str += "\n" + cmd_output_spacer + output + "\n"
                write_str += cmd_output_spacer + spacer
                filehandle.write(write_str)
            write_str = "*" * 23 + " End of File " + "*" * 24
            filehandle.write(write_str)

def init_next_row_cache(wb_obj, sheets):
    cache = {}
    for sheet_name in sheets:
        ws = wb_obj[sheet_name]
        # If column A is your “has data” column:
        cache[sheet_name] = next_available_row(ws, "A")  # one scan per sheet
    return cache

def gather_results_to_wb(wb_obj, net_dev, key_map, headers_key, next_row_cache):
    """
    Writes the Simple show commands to the work book based on the
    key_map information.
    """
    #headers_key = map_headers(wb_obj)
    show_results = net_dev.show_for_xls
    d_parse = net_dev.parse_method
    hostname = net_dev.hostname
    for setting, list_of_shows in show_results.items():
        if isinstance(list_of_shows, list):
            sheet = list(key_map[d_parse][setting].keys())[0]
            c_mapper = key_map[d_parse][setting][sheet]
            for value in list_of_shows:
                #row = next_available_row(wb_obj[sheet])
                row = next_row_cache[sheet]
                next_row_cache[sheet] += 1
                rw_cell(wb_obj[sheet], row, 1, True, hostname)
                for key, c_name in c_mapper.items():
                    if key in value.keys():
                        col = headers_key[sheet][c_name]
                        wr_val = value[key]
                        rw_cell(wb_obj[sheet], row, col, True, wr_val)
        else:
            print("Error", list_of_shows)

def post_process_lldp_sheet_with_wtp_versions(wb_obj, headers_key, all_devices):
    """
    After ALL devices have been gathered/written:
    - scan the LLDP sheet
    - for each Fortinet device's fortinet_wtp_status entries:
        find matching LLDP row(s) and append WTP software-version to System Description
    """

    print("Running Fortinet wtp-status to LLDP post-processing")

    ws = wb_obj["LLDP"]

    # ---- resolve LLDP sheet columns (robust to header wording) ----
    hdr = headers_key["LLDP"]  # mapping: header_text -> col_index

    def col_for(*candidates):
        # candidates are header names you might be using in the sheet
        for c in candidates:
            if c in hdr:
                return hdr[c]
        # fallback: try "contains" match
        for c in candidates:
            cl = c.lower()
            for k, v in hdr.items():
                if cl in k.lower():
                    return v
        return None

    COL_LOCAL_HOST = 1  # you always write hostname into col 1 in gather_results_to_wb()
    COL_REMOTE_HOST = col_for("System Name", "Remote Host", "Sys Name", "System name")
    COL_MGMT_IP     = col_for("Management IP", "Mgmt IP", "IP", "ip")
    COL_PORT_ID     = col_for("Port ID", "Remote Port", "Port id")
    COL_CHASSIS_ID  = col_for("Chassis ID", "Chassis id")
    COL_SOFTWARE     = col_for("Software")

    if not COL_SOFTWARE:
        raise RuntimeError("LLDP post-process: could not find Remote Description column in LLDP headers")

    # ---- build an index of LLDP rows by local hostname (col A) ----
    # We only need quick filtering by local device (the switch).
    rows_by_local = {}
    max_row = ws.max_row

    for r in range(2, max_row + 1):  # skip header row
        local = ws.cell(r, COL_LOCAL_HOST).value
        if not local:
            continue
        local_key = str(local).strip().lower()
        rows_by_local.setdefault(local_key, []).append(r)

    def norm_host(s: str) -> str:
        s = (s or "").strip().lower()
        return s.split(".")[0] if s else ""

    def norm_mac(s: str) -> str:
        return (s or "").strip().lower()

    def append_once(existing: str, tag: str) -> str:
        existing = (existing or "").strip()
        if not tag:
            return existing
        return existing if tag in existing else (existing + (" | " if existing else "") + tag)

    # ---- for each WTP record, find and update matching LLDP row(s) ----
    for fg in all_devices:
        if getattr(fg, "parse_method", "") != "fortinet":
            continue

        wtps = getattr(fg, "fortinet_wtp_status", None) or []
        if not wtps:
            continue

        for w in wtps:
            w = {k.lower(): v for k, v in (w or {}).items()}

            wtp_name = (w.get("wtp_name") or "").strip()
            wtp_sw   = (w.get("software_version") or "").strip()
            wtp_ip   = (w.get("local_ipv4_addr") or w.get("mgmt_ip") or "").strip()
            board_mac = norm_mac(w.get("board_mac"))
            sw_host = norm_host(w.get("lldp_sys_name"))  # switch hostname from the AP's LLDP block
            sw_host_fqdn = (w.get("lldp_sys_name") or "").strip().lower()

            if not wtp_sw:
                continue

            # Candidate local device keys in sheet
            candidate_locals = {sw_host, sw_host_fqdn}
            candidate_rows = []
            for lk in candidate_locals:
                if lk and lk in rows_by_local:
                    candidate_rows.extend(rows_by_local[lk])

            if not candidate_rows:
                continue

            # Try to match row by: MGMT IP, OR Remote Port (board-mac), OR Remote Host (WTP name)
            for r in candidate_rows:
                row_remote_host = str(ws.cell(r, COL_REMOTE_HOST).value or "").strip() if COL_REMOTE_HOST else ""
                row_mgmt_ip     = str(ws.cell(r, COL_MGMT_IP).value or "").strip() if COL_MGMT_IP else ""
                row_port_id     = str(ws.cell(r, COL_PORT_ID).value or "").strip() if COL_PORT_ID else ""
                row_chassis     = norm_mac(ws.cell(r, COL_CHASSIS_ID).value or "") if COL_CHASSIS_ID else ""

                host_match = bool(wtp_name and row_remote_host and row_remote_host.lower() == wtp_name.lower())
                ip_match   = bool(wtp_ip and row_mgmt_ip and row_mgmt_ip == wtp_ip)
                port_match = False

                if board_mac:
                    port_match = (norm_mac(row_port_id) == board_mac) or (row_chassis == board_mac)

                if host_match or ip_match or port_match:
                    tag = f"WTP_SW={wtp_sw}"
                    cur = ws.cell(r, COL_SOFTWARE).value
                    ws.cell(r, COL_SOFTWARE).value = append_once(str(cur or ""), tag)
                    break

def add_err_msgs_to_wb(net_dev, wb_obj):
    """
    Adds all the Registered Errors to the WorkSheet.
    """
    sheet_obj = wb_obj["Errors"]
    err_msgs = net_dev.error_msgs
    row = next_available_row(sheet_obj)
    hostname = net_dev.hostname
    if not hostname:
        hostname = net_dev.host
    for i, msg in enumerate(err_msgs):
        row += i
        rw_cell(sheet_obj, row, 1, True, hostname)
        rw_cell(sheet_obj, row, 2, True, msg[1])
        rw_cell(sheet_obj, row, 3, True, msg[0])
