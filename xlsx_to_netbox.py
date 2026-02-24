# 

#!/usr/bin/env python3

# conda install ahuff::pynetbox
# conda install main::openpyxl

#!/usr/bin/env python3
"""xlsx_to_netbox.py

Optimized XLSX -> NetBox importer (devices, interfaces, IPs, CDP/LLDP custom field)

Behavior (per our conversation):
  - Main sheet:
      Row 7 header; data from row 8
      A = Host IP (no mask) used to derive mgmt_ip
      B = status source: Yes/Completed -> active else offline
      C = platform string; if contains "cisco_ios" -> manufacturer Cisco else Unknown
      J = device name
      K = vendor model string (device_type.model); create if missing
  - Interfaces sheet:
      A = Hostname (matches Main!J)
      B = Interface name
      D = Description
      E = Human interface type string (mapped to NetBox interface type slug)
      G = Link: up/down => enabled True, anything else False
      J = Mode: access/trunk/tagged -> NetBox access/tagged
      K = Access VLAN like "VLAN50 (50)" (VID extracted)
      N = Trunk allowed VLANs string like "10,20,30-40"
      P = IP Add with mask
      Q = MTU
  - CDP/LLDP sheets:
      CDP headers: Hostname, Remote Host, MGMT IP, Platform, Software
      LLDP headers: Hostname, Remote Host, MGMT IP, Software (Platform not required; stored blank)
  - NetBox:
      Site fixed to DEFAULT_SITE_NAME
      Role: "Access Switch" but only set on UPDATE if device role is currently unset.
      Device primary_ip4 is set AFTER interface IP assignment.
      custom_fields["cdp_neighbors_json"] stored as a serialized JSON string.

Performance optimizations implemented:
  1) Prefetch interfaces once per device (no per-row .get())
  2) Bulk-create missing interfaces per device
  3) Run-level caches for role/manufacturer/device_type/VLAN/IP
  4) Only PATCH (save) when something actually changed
  5) Enable PyNetBox threading for faster multi-page fetches

Refs:
  - PyNetBox bulk create by passing list of dicts to .create():
    https://pynetbox.readthedocs.io/en/stable/endpoint.html
  - PyNetBox threading for faster .filter()/.all():
    https://netboxlabs.com/docs/sdks/pynetbox/advanced/
"""

from __future__ import annotations

import argparse
import ipaddress
import json
import os
import re
from typing import Any, Dict, List, Optional, Tuple
from pynetbox.core.query import RequestError
from datetime import datetime
from pathlib import Path
from time import perf_counter

import openpyxl
import pynetbox


DEFAULT_SITE_NAME = "SCDS IT Dept."
DEFAULT_MANUFACTURER_CISCO = "Cisco"
DEFAULT_DEVICE_ROLE_NAME = "Access Switch"
VLAN_BEHAVIOR = "strict"


# -----------------------------
# Run-level caches (big perf win)
# -----------------------------
ROLE_CACHE: Dict[str, Any] = {}
MFG_CACHE: Dict[str, Any] = {}
DEVTYPE_CACHE: Dict[Tuple[int, str], Any] = {}        # (manufacturer_id, model)
VLAN_CACHE: Dict[Tuple[int, int], Any] = {}           # (site_id, vid)
VLAN_GROUP_CACHE = {}
IP_CACHE_BY_ADDRESS: Dict[str, Any] = {}              # "1.2.3.4/24" -> ip obj
IP_CACHE_BY_HOST: Dict[str, Any] = {}                 # "1.2.3.4" -> ip obj

class VlanScopeError(RuntimeError):
    """Raised when a VLAN VID exists but is not in the expected Global group/scope."""
    pass

# -----------------------------
# Inventory sheet -> Interface type override
# -----------------------------
def _norm_type_key(v: Any) -> str:
    """Normalize free-text optics/media strings for reliable matching."""
    if v is None:
        return ""
    s = str(v).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s

# Map Inventory!Description -> NetBox interface 'type' slug
# (Keys are normalized with _norm_type_key)
INVENTORY_DESC_TO_NBTYPE: Dict[str, str] = {
    _norm_type_key("SFP-10GBase-SR"): "10gbase-x-sfpp",      # display: SFP+ (10GE)
    _norm_type_key("SFP-10Gbase-SR"): "10gbase-x-sfpp",
    _norm_type_key("SFP-10Gbase-LR"): "10gbase-x-sfpp",
    _norm_type_key("4500X-16 10GE (SFP+)"): "10gbase-x-sfpp",
    _norm_type_key("10GE SFP+"): "10gbase-x-sfpp",

    _norm_type_key("1000BaseSX SFP"): "1000base-sx",
    _norm_type_key("1000BaseSX"): "1000base-sx",
    _norm_type_key("1000BaseLH"): "1000base-lx10",          # display: 1000BASE-LX10/LH
    _norm_type_key("1000BaseLX SFP"): "1000base-lx",
    _norm_type_key("SFP-10GBase-LR"): "10gbase-lr",         # display: 10GBASE-LR
}

# -----------------------------
# Helpers
# -----------------------------
def slugify(s: str) -> str:
    s = s.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-{2,}", "-", s).strip("-")
    return s or "item"

def map_device_status(v: Any) -> str:
    s = "" if v is None else str(v).strip().lower()
    return "active" if s in ("yes", "completed") else "offline"

def parse_vlan_id(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, int):
        return v
    s = str(v).strip()
    if not s:
        return None
    m = re.search(r"\((\d+)\)", s)
    if m:
        return int(m.group(1))
    m = re.search(r"\b(\d+)\b", s)
    if m:
        return int(m.group(1))
    return None

def expand_vlan_list(s: str) -> List[int]:
    vlans: List[int] = []
    if not s:
        return vlans
    for part in str(s).split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            try:
                vlans.extend(range(int(a), int(b) + 1))
            except ValueError:
                continue
        else:
            try:
                vlans.append(int(part))
            except ValueError:
                continue
    return sorted(set(vlans))

def map_interface_type(human: Any, inventory_desc: Any = None) -> str:
    """Return NetBox interface type slug.

    Precedence:
      1) Inventory sheet (authoritative optics/media description per-port)
      2) Interfaces sheet "Type" column exact mapping (common values)
      3) Heuristics (fallback)
    """
    # 1) Inventory override (most accurate for physical media)
    inv_key = _norm_type_key(inventory_desc)
    if inv_key and inv_key in INVENTORY_DESC_TO_NBTYPE:
        return INVENTORY_DESC_TO_NBTYPE[inv_key]

    # 2) Interfaces sheet explicit mapping
    human_key = _norm_type_key(human)
    INTERFACES_TYPE_TO_NBTYPE: Dict[str, str] = {
        _norm_type_key("Gigabit Ethernet"): "1000base-t",
        _norm_type_key("Ten Gigabit Ethernet"): "10gbase-t",
        _norm_type_key("Ten Gigabit Ethernet Port"): "10gbase-t",
        _norm_type_key("PowerPC FastEthernet"): "100base-tx",
        _norm_type_key("RP management port"): "100base-tx",
        _norm_type_key("EtherChannel"): "lag",
        # SVIs are virtual interfaces in NetBox
        _norm_type_key("EtherSVI"): "virtual",
        _norm_type_key("Ethernet SVI"): "virtual",
    }
    if human_key and human_key in INTERFACES_TYPE_TO_NBTYPE:
        return INTERFACES_TYPE_TO_NBTYPE[human_key]

    # 3) Heuristics
    s = "" if human is None else str(human).strip()
    u = s.upper()

    if "SFP+" in u or "SFPP" in u or "10GBASE" in u or "10GB" in u:
        return "10gbase-x-sfpp"
    if "10G" in u:
        return "10gbase-t"
    if "1000" in u and ("SFP" in u or "BASESX" in u or "BASELX" in u or "BASE-X" in u):
        return "1000base-x-sfp"
    if "1000BASE" in u or "1GE" in u or re.search(r"\b1g\b", u):
        return "1000base-t"
    if "100BASE" in u or "FAST" in u:
        return "100base-tx"
    if "SVI" in u or "VLAN" in u:
        return "virtual"

    return "other"

def map_mode(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip().lower()
    if not s:
        return None
    if "access" in s:
        return "access"
    if "trunk" in s or "tagged" in s:
        return "tagged"
    if "routed" in s:
        return None
    return None

def link_to_enabled(v: Any) -> bool:
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in ("up", "down")

def is_cisco_platform(platform_str: Any) -> bool:
    if platform_str is None:
        return False
    return "cisco_ios" in str(platform_str).lower()

def resolve_mgmt_ip_with_mask(main_host: str, intfs_for_host: List[Dict[str, Any]]) -> Tuple[str, bool]:
    """Return (mgmt_ip_cidr, matched).

    Finds an interface IP (with mask) whose host matches Main host.
    If none found, returns MainHost/24 and matched=False.
    """
    host = str(main_host).strip()
    try:
        host_ip = ipaddress.ip_address(host)
    except ValueError:
        if "/" in host:
            return host, True
        return f"{host}/24", False

    for intf in intfs_for_host:
        ip_val = intf.get("ip_address")
        if not ip_val:
            continue
        ip_s = str(ip_val).strip()
        if not ip_s:
            continue
        try:
            iface = ipaddress.ip_interface(ip_s)
        except ValueError:
            continue
        if iface.ip == host_ip:
            return str(iface), True

    return f"{host}/24", False

def prefer_mgmt_interface(intfs: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not intfs:
        return None
    # Prefer Vlan1 then Mgmt0 then first
    for wanted in ("vlan1", "mgmt0", "management", "vlan", "loopback0"):
        for r in intfs:
            if str(r.get("name", "")).strip().lower() == wanted:
                return r
    return intfs[0]

def sheet_headers(ws) -> Dict[str, int]:
    hdr = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        name = str(v).strip()
        if name:
            hdr[name] = col
    return hdr

def pick_best_ip(ip_objs, desired_iface=None):
    """
    Prefer:
      1) exact address match (incl mask) if desired_iface provided
      2) global table (vrf is None)
      3) otherwise first result
    """
    ip_objs = list(ip_objs)
    if not ip_objs:
        return None

    if desired_iface is not None:
        desired_addr = str(desired_iface)
        for o in ip_objs:
            if getattr(o, "address", None) == desired_addr:
                return o

    for o in ip_objs:
        if getattr(o, "vrf", None) in (None, ""):
            return o

    return ip_objs[0]


# Neighbor helpers go here:
def _norm(s: Any) -> str:
    return "" if s is None else str(s).strip()

def _key(host: Any, port: Any) -> Tuple[str, str]:
    return (_norm(host), _norm(port))

def build_neighbors_from_cdp(ws_cdp) -> Dict[Tuple[str, str], str]:
    """
    Return dict[(hostname, local_port)] = remote_host

    Match rule:
      - CDP.Hostname + CDP.Local Port  <->  Interfaces.Hostname + Interfaces.Interface

    Inclusion rule (per your spec):
      - requires Hostname, Local Port, Remote Host
    """
    hdr = sheet_headers(ws_cdp)
    required = ["Hostname", "Local Port", "Remote Host"]
    for k in required:
        if k not in hdr:
            raise RuntimeError(f"Sheet '{ws_cdp.title}' missing required header '{k}'")

    out: Dict[Tuple[str, str], str] = {}
    for r in range(2, ws_cdp.max_row + 1):
        host = ws_cdp.cell(row=r, column=hdr["Hostname"]).value
        local = ws_cdp.cell(row=r, column=hdr["Local Port"]).value
        remote = ws_cdp.cell(row=r, column=hdr["Remote Host"]).value

        if not _norm(host) or not _norm(local) or not _norm(remote):
            continue

        out[_key(host, local)] = _norm(remote)

    return out

def build_neighbors_from_lldp(ws_lldp) -> Dict[Tuple[str, str], str]:
    """
    Return dict[(hostname, local_port)] = description_string

    Match rule:
      - LLDP.Hostname + LLDP.Local Port  <->  Interfaces.Hostname + Interfaces.Short IF

    Inclusion rule (per your spec):
      - ONLY set description from LLDP where:
          Local Port, Remote Host, Remote Description, and Software have values
    """
    hdr = sheet_headers(ws_lldp)
    required = ["Hostname", "Local Port", "Remote Host", "Remote Description", "Software"]
    for k in required:
        if k not in hdr:
            raise RuntimeError(f"Sheet '{ws_lldp.title}' missing required header '{k}'")

    out: Dict[Tuple[str, str], str] = {}
    for r in range(2, ws_lldp.max_row + 1):
        host = ws_lldp.cell(row=r, column=hdr["Hostname"]).value
        local = ws_lldp.cell(row=r, column=hdr["Local Port"]).value
        remote = ws_lldp.cell(row=r, column=hdr["Remote Host"]).value
        rdesc = ws_lldp.cell(row=r, column=hdr["Remote Description"]).value
        software = ws_lldp.cell(row=r, column=hdr["Software"]).value

        # strict inclusion rule
        if not (_norm(host) and _norm(local) and _norm(remote) and _norm(rdesc) and _norm(software)):
            continue

        # what you want in Interfaces.description
        out[_key(host, local)] = f"{_norm(remote)} ({_norm(rdesc)})"

    return out

def build_port_mgmtip_from_cdp(ws_cdp) -> Dict[Tuple[str, str], str]:
    """
    Return dict[(hostname, local_port)] = mgmt_ip

    Match rule:
      - CDP.Hostname + CDP.Local Port  <->  Interfaces.Hostname + Interfaces.Interface

    Inclusion:
      - requires Hostname, Local Port, MGMT IP (Remote Host not required for IP assignment)
    """
    hdr = sheet_headers(ws_cdp)
    required = ["Hostname", "Local Port", "MGMT IP"]
    for k in required:
        if k not in hdr:
            raise RuntimeError(f"Sheet '{ws_cdp.title}' missing required header '{k}'")

    out: Dict[Tuple[str, str], str] = {}
    for r in range(2, ws_cdp.max_row + 1):
        host = ws_cdp.cell(row=r, column=hdr["Hostname"]).value
        local = ws_cdp.cell(row=r, column=hdr["Local Port"]).value
        mgmt = ws_cdp.cell(row=r, column=hdr["MGMT IP"]).value

        if not _norm(host) or not _norm(local) or not _norm(mgmt):
            continue

        out[_key(host, local)] = _norm(mgmt)

    return out

def build_port_mgmtip_from_lldp(ws_lldp) -> Dict[Tuple[str, str], str]:
    """
    Return dict[(hostname, local_port)] = mgmt_ip

    Match rule:
      - LLDP.Hostname + LLDP.Local Port  <->  Interfaces.Hostname + Interfaces.Short IF

    Inclusion:
      - requires Hostname, Local Port, MGMT IP
    """
    hdr = sheet_headers(ws_lldp)
    required = ["Hostname", "Local Port", "MGMT IP"]
    for k in required:
        if k not in hdr:
            raise RuntimeError(f"Sheet '{ws_lldp.title}' missing required header '{k}'")

    out: Dict[Tuple[str, str], str] = {}
    for r in range(2, ws_lldp.max_row + 1):
        host = ws_lldp.cell(row=r, column=hdr["Hostname"]).value
        local = ws_lldp.cell(row=r, column=hdr["Local Port"]).value
        mgmt = ws_lldp.cell(row=r, column=hdr["MGMT IP"]).value

        if not _norm(host) or not _norm(local) or not _norm(mgmt):
            continue

        out[_key(host, local)] = _norm(mgmt)

    return out

def normalize_device_name(name: str) -> str:
    """Strip FQDN and whitespace: 'SCDS_XYZ.domain.tld' -> 'SCDS_XYZ'."""
    s = (name or "").strip()
    if not s:
        return s
    return s.split(".", 1)[0]

def normalize_mac(s: Any) -> str:
    """Normalize mac strings like '0c85.255c.0e00' or '0C:85:25:5C:0E:00' -> '0c85255c0e00'."""
    if s is None:
        return ""
    return re.sub(r"[^0-9a-fA-F]", "", str(s)).lower()

def expand_cisco_ifname(port: str) -> str:
    """
    Expand common Cisco abbreviations:
      Te1/1 -> TenGigabitEthernet1/1
      Gi1/0/1 -> GigabitEthernet1/0/1
      Fa0/1 -> FastEthernet0/1
      Po1 -> Port-channel1
    If already long-form, return as-is.
    """
    p = (port or "").strip()
    if not p:
        return p

    # Already long-ish
    if p.lower().startswith(("tengigabitethernet", "gigabitethernet", "fastethernet", "port-channel", "ethernet")):
        return p

    m = re.match(r"^(Te|Gi|Fa|Po|Eth)\s*(.+)$", p, re.IGNORECASE)
    if not m:
        return p

    pref, rest = m.group(1).lower(), m.group(2).strip()
    if pref == "te":
        return f"TenGigabitEthernet{rest}"
    if pref == "gi":
        return f"GigabitEthernet{rest}"
    if pref == "fa":
        return f"FastEthernet{rest}"
    if pref == "po":
        return f"Port-channel{rest}"
    if pref == "eth":
        return f"Ethernet{rest}"
    return p

def ensure_cable_between_interfaces_orig(nb, a_intf, b_intf, commit: bool) -> None:
    """
    Create a cable between a_intf and b_intf if neither is already cabled.
    Only creates if both interface objects exist (callers ensure that).
    """
    if getattr(a_intf, "cable", None) or getattr(b_intf, "cable", None):
        return

    if not commit:
        print(f"[DRY] Would cable {a_intf.device.name}:{a_intf.name} <-> {b_intf.device.name}:{b_intf.name}")
        return

    nb.dcim.cables.create({
        "termination_a_type": "dcim.interface",
        "termination_a_id": a_intf.id,
        "termination_b_type": "dcim.interface",
        "termination_b_id": b_intf.id,
        "status": "connected",
    })

def ensure_cable_between_interfaces(nb, a_intf, b_intf, commit: bool) -> None:
    """
    Create a cable between a_intf and b_intf.
    - Uses NetBox v3.3+ cable API (a_terminations/b_terminations)
    - Never raises RequestError (so you don't skip the whole device)
    - Only creates if both ends exist and neither end is already cabled
    """
    if not a_intf or not b_intf:
        return

    a_id = getattr(a_intf, "id", None)
    b_id = getattr(b_intf, "id", None)
    if not a_id or not b_id:
        return

    # If either end already has a cable, don't touch it
    if getattr(a_intf, "cable", None) or getattr(b_intf, "cable", None):
        return

    payload = {
        "status": "connected",
        "a_terminations": [{"object_type": "dcim.interface", "object_id": a_id}],
        "b_terminations": [{"object_type": "dcim.interface", "object_id": b_id}],
    }

    if not commit:
        print(f"[DRY] Would cable {a_intf.device.name}:{a_intf.name} <-> {b_intf.device.name}:{b_intf.name}")
        return

    try:
        nb.dcim.cables.create(payload)
    except RequestError as e:
        print(
            f"[WARN] Cable create failed for {a_intf.device.name}:{a_intf.name} <-> "
            f"{b_intf.device.name}:{b_intf.name}: {e}"
        )
        return

def build_cdp_links(ws_cdp) -> Dict[Tuple[str, str], Tuple[str, str]]:
    """
    (local_host, local_port) -> (remote_host, remote_port)
    """
    hdr = sheet_headers(ws_cdp)
    required = ["Hostname", "Local Port", "Remote Host", "Remote Port"]
    for k in required:
        if k not in hdr:
            raise RuntimeError(f"Sheet '{ws_cdp.title}' missing required header '{k}'")

    out: Dict[Tuple[str, str], Tuple[str, str]] = {}
    for r in range(2, ws_cdp.max_row + 1):
        lh = normalize_device_name(_norm(ws_cdp.cell(row=r, column=hdr["Hostname"]).value))
        lp = _norm(ws_cdp.cell(row=r, column=hdr["Local Port"]).value)
        rh = normalize_device_name(_norm(ws_cdp.cell(row=r, column=hdr["Remote Host"]).value))
        rp = _norm(ws_cdp.cell(row=r, column=hdr["Remote Port"]).value)
        if not (lh and lp and rh and rp):
            continue
        out[(lh, lp)] = (rh, rp)
    return out


def build_lldp_links(ws_lldp, interfaces_by_host: Dict[str, List[Dict[str, Any]]]) -> Dict[Tuple[str, str], Tuple[str, str]]:
    """
    (local_host, local_port) -> (remote_host, remote_port)

    If Local Port is blank, use Chassis ID as a MAC pointer:
      - chassis_id == Interfaces!MAC Add for the local interface
      - then use that interface's 'short_if' (preferred) or full 'name'
    """
    hdr = sheet_headers(ws_lldp)
    required = ["Hostname", "Chassis ID", "Local Port", "Remote Host", "Remote Port"]
    for k in required:
        if k not in hdr:
            raise RuntimeError(f"Sheet '{ws_lldp.title}' missing required header '{k}'")

    # Build per-host MAC->(short_if/full_name)
    mac_index: Dict[Tuple[str, str], str] = {}
    for host, rows in interfaces_by_host.items():
        h = normalize_device_name(host)
        for row in rows:
            m = normalize_mac(row.get("mac"))
            if not m:
                continue
            # Prefer short_if because LLDP often uses Te1/1 style
            mac_index[(h, m)] = row.get("short_if") or row.get("name") or ""

    out: Dict[Tuple[str, str], Tuple[str, str]] = {}
    for r in range(2, ws_lldp.max_row + 1):
        lh = normalize_device_name(_norm(ws_lldp.cell(row=r, column=hdr["Hostname"]).value))
        chassis = _norm(ws_lldp.cell(row=r, column=hdr["Chassis ID"]).value)
        lp = _norm(ws_lldp.cell(row=r, column=hdr["Local Port"]).value)
        rh = normalize_device_name(_norm(ws_lldp.cell(row=r, column=hdr["Remote Host"]).value))
        rp = _norm(ws_lldp.cell(row=r, column=hdr["Remote Port"]).value)

        if not (lh and rh and rp):
            continue

        # Local Port missing: resolve via chassis-id MAC -> Interfaces.MAC Add
        if not lp and chassis:
            lp = mac_index.get((lh, normalize_mac(chassis)), "")

        if not lp:
            continue

        out[(lh, lp)] = (rh, rp)

    return out


# -----------------------------
# NetBox helpers (cached)
# -----------------------------
def nb_get_or_create_site(nb, site_name: str) -> Any:
    site = nb.dcim.sites.get(name=site_name)
    if site:
        return site
    raise RuntimeError(f"Site '{site_name}' not found in NetBox (expected to exist).")

def nb_get_or_create_manufacturer(nb, name: str, commit: bool) -> Any:
    if name in MFG_CACHE:
        return MFG_CACHE[name]
    m = nb.dcim.manufacturers.get(name=name)
    if m:
        MFG_CACHE[name] = m
        return m
    if not commit:
        obj = type("Obj", (), {"id": -1, "name": name})
        MFG_CACHE[name] = obj
        return obj
    m = nb.dcim.manufacturers.create(name=name, slug=slugify(name))
    MFG_CACHE[name] = m
    return m

def nb_get_or_create_role(nb, name: str, commit: bool) -> Any:
    if name in ROLE_CACHE:
        return ROLE_CACHE[name]
    r = nb.dcim.device_roles.get(name=name)
    if r:
        ROLE_CACHE[name] = r
        return r
    if not commit:
        obj = type("Obj", (), {"id": -1, "name": name})
        ROLE_CACHE[name] = obj
        return obj
    r = nb.dcim.device_roles.create(name=name, slug=slugify(name))
    ROLE_CACHE[name] = r
    return r

def nb_get_or_create_device_type(nb, manufacturer_id: int, model: str, commit: bool) -> Any:
    key = (manufacturer_id, model)
    if key in DEVTYPE_CACHE:
        return DEVTYPE_CACHE[key]
    dt = nb.dcim.device_types.get(model=model)
    if dt:
        DEVTYPE_CACHE[key] = dt
        return dt
    payload = {"manufacturer": manufacturer_id, "model": model, "slug": slugify(model)}
    if not commit:
        obj = type("Obj", (), {"id": -1, "model": model})
        DEVTYPE_CACHE[key] = obj
        return obj
    dt = nb.dcim.device_types.create(**payload)
    DEVTYPE_CACHE[key] = dt
    return dt

def nb_get_or_create_vlan_group_orig(nb, name: str = "Global", commit: bool = True):
    """
    Ensure VLAN group exists and return it.
    """
    vg = nb.ipam.vlan_groups.get(name=name)
    if vg:
        return vg

    if not commit:
        # dry run: return None or a small stub if your code expects .id
        return None

    return nb.ipam.vlan_groups.create({"name": name, "slug": name.lower().replace(" ", "-")})

def nb_get_or_create_vlan_group(nb, name: str, commit: bool):
    if name in VLAN_GROUP_CACHE:
        return VLAN_GROUP_CACHE[name]

    vg = nb.ipam.vlan_groups.get(name=name)
    if vg:
        VLAN_GROUP_CACHE[name] = vg
        return vg

    if not commit:
        obj = type("Obj", (), {"id": -1, "name": name})
        VLAN_GROUP_CACHE[name] = obj
        return obj

    vg = nb.ipam.vlan_groups.create(name=name, slug=slugify(name))
    VLAN_GROUP_CACHE[name] = vg
    return vg

def nb_get_or_create_vlan_orig(nb, site_id: int, vid: int, commit: bool) -> Any:
    key = (site_id, vid)
    if key in VLAN_CACHE:
        return VLAN_CACHE[key]
    vlan = nb.ipam.vlans.get(site_id=site_id, vid=vid)
    if vlan:
        VLAN_CACHE[key] = vlan
        return vlan
    payload = {"site": None, "vid": vid, "name": f"VLAN{vid}", "status": "active"}
    if not commit:
        obj = type("Obj", (), {"id": -1, "vid": vid})
        VLAN_CACHE[key] = obj
        return obj
    vlan = nb.ipam.vlans.create(**payload)
    VLAN_CACHE[key] = vlan
    return vlan


# global flag set in main: VLAN_BEHAVIOR = "strict" or "normalize"
# requires: nb_get_or_create_vlan_group(), VlanScopeError, VLAN_CACHE
def nb_get_or_create_vlan(nb, site_id: int, vid: int, commit: bool) -> Any:
    """
    Global VLAN logic; only difference between modes is how we handle VID that exists outside Global.

    Both modes:
      - If VID exists in Global:
          * if duplicates, use lowest-id canonical and continue (warn)
      - If VID does not exist anywhere: create it in Global (site=None)

    strict:
      - If VID exists outside Global (but not in Global): raise VlanScopeError

    normalize:
      - If VID exists outside Global (but not in Global): move it into Global and clear site (site=None)
    """
    global VLAN_BEHAVIOR

    vid = int(vid)
    cache_key = ("Global", vid)
    if cache_key in VLAN_CACHE:
        return VLAN_CACHE[cache_key]

    vg = nb_get_or_create_vlan_group(nb, "Global", commit)
    vg_id = getattr(vg, "id", vg)

    # 1) Look in Global by VID
    global_matches = list(nb.ipam.vlans.filter(group_id=vg_id, vid=vid))
    if global_matches:
        # duplicates -> canonicalize (lowest id) and continue
        global_matches = sorted(global_matches, key=lambda v: v.id)
        canonical = global_matches[0]
        if len(global_matches) > 1:
            print(
                f"[WARN] Duplicate VLANs in Global for VID {vid}: {[v.id for v in global_matches]}. "
                f"Using canonical id={canonical.id}."
            )
        VLAN_CACHE[cache_key] = canonical
        return canonical

    # 2) Not in Global. Does it exist elsewhere?
    elsewhere = list(nb.ipam.vlans.filter(vid=vid))
    if elsewhere:
        # pick lowest id elsewhere (deterministic)
        elsewhere = sorted(elsewhere, key=lambda v: v.id)
        vlan = elsewhere[0]

        grp = getattr(getattr(vlan, "group", None), "name", None)
        site = getattr(getattr(vlan, "site", None), "name", None)

        if VLAN_BEHAVIOR == "strict":
            raise VlanScopeError(
                f"VID {vid} exists outside Global (vlan_id={vlan.id}, group={grp!r}, site={site!r}); "
                f"strict mode refuses to move it."
            )

        # normalize: move into Global + clear site
        if not commit:
            obj = type("Obj", (), {"id": vlan.id, "vid": vid})
            VLAN_CACHE[cache_key] = obj
            return obj

        vlan.update({"group": vg_id, "site": None})
        VLAN_CACHE[cache_key] = vlan
        return vlan

    # 3) Doesn't exist anywhere -> create in Global
    payload = {"site": None, "group": vg_id, "vid": vid, "name": f"VLAN{vid}", "status": "active"}

    if not commit:
        obj = type("Obj", (), {"id": -1, "vid": vid})
        VLAN_CACHE[cache_key] = obj
        return obj

    vlan = nb.ipam.vlans.create(**payload)
    VLAN_CACHE[cache_key] = vlan
    return vlan

def nb_get_or_create_ip(nb, address: str, commit: bool) -> Any:
    """
    Get or create an IPAddress record, handling:
      - host-uniqueness behavior (same host different masks collide in global table)
      - multiple matches (VRFs, duplicates) by using filter() + pick_best_ip()
      - safe mask updates only when there is no exact desired object already

    Returns a pynetbox IP record (or dry-run stub).
    """
    addr_s = str(address).strip()
    if not addr_s:
        raise ValueError("nb_get_or_create_ip: empty address")

    # Normalize to an ip_interface (ensure mask)
    desired = ipaddress.ip_interface(addr_s) if "/" in addr_s else ipaddress.ip_interface(f"{addr_s}/32")
    desired_str = str(desired)
    host_str = str(desired.ip)

    # 0) Cache hits
    if desired_str in IP_CACHE_BY_ADDRESS:
        return IP_CACHE_BY_ADDRESS[desired_str]

    # If we have a host cached, prefer an exact desired CIDR if it exists,
    # otherwise return the cached host object WITHOUT mutating its mask here.
    if host_str in IP_CACHE_BY_HOST:
        exact = pick_best_ip(nb.ipam.ip_addresses.filter(address=desired_str), desired)
        if exact:
            IP_CACHE_BY_ADDRESS[desired_str] = exact
            IP_CACHE_BY_HOST[host_str] = exact
            return exact
        return IP_CACHE_BY_HOST[host_str]

    # 1) Exact CIDR match (may be >1 record; use filter + pick)
    exact = pick_best_ip(nb.ipam.ip_addresses.filter(address=desired_str), desired)
    if exact:
        IP_CACHE_BY_ADDRESS[desired_str] = exact
        IP_CACHE_BY_HOST[host_str] = exact
        return exact

    # 2) Host match (find any record whose host == desired.ip)
    candidates = list(nb.ipam.ip_addresses.filter(q=host_str))
    matched = None
    for c in candidates:
        try:
            ex = ipaddress.ip_interface(c.address)
        except Exception:
            continue
        if ex.ip == desired.ip:
            matched = c
            break

    if matched:
        # Cache what we found
        try:
            ex = ipaddress.ip_interface(matched.address)
            IP_CACHE_BY_ADDRESS[str(ex)] = matched
        except Exception:
            pass
        IP_CACHE_BY_HOST[host_str] = matched

        # If mask differs, only update if there's definitely no exact desired object.
        # (We already checked exact above, but check again for safety/races.)
        if str(ipaddress.ip_interface(matched.address)) != desired_str:
            exact2 = pick_best_ip(nb.ipam.ip_addresses.filter(address=desired_str), desired)
            if exact2:
                IP_CACHE_BY_ADDRESS[desired_str] = exact2
                IP_CACHE_BY_HOST[host_str] = exact2
                return exact2

            if commit:
                try:
                    matched.address = desired_str
                    matched.save()
                except Exception:
                    # If NetBox refuses due to uniqueness/duplicates, keep the existing record as-is.
                    pass

        return matched

    # 3) Create new (bulk ensure uses different path; this is for point lookups)
    payload = {"address": desired_str, "status": "active"}
    if not commit:
        obj = type("Obj", (), {"id": -1, "address": desired_str})
        IP_CACHE_BY_ADDRESS[desired_str] = obj
        IP_CACHE_BY_HOST[host_str] = obj
        return obj

    created = nb.ipam.ip_addresses.create(**payload)
    IP_CACHE_BY_ADDRESS[desired_str] = created
    IP_CACHE_BY_HOST[host_str] = created
    return created

def ensure_ips_exist_bulk(nb, ip_addrs: List[str], commit: bool) -> None:
    """Bulk create IPs that don't exist.

    Key fixes:
      - Deduplicate within this call (mgmt_ip often equals an interface IP)
      - "Reserve" entries in cache while building 'needed' so duplicates in the same loop
        don't get appended twice
      - Use the return value of .create(list_of_dicts) to prime cache directly
    """
    needed: List[Dict[str, str]] = []
    seen_desired: set[str] = set()

    for addr in ip_addrs:
        if not addr:
            continue
        addr_s = str(addr).strip()
        if not addr_s:
            continue

        desired = ipaddress.ip_interface(addr_s) if "/" in addr_s else ipaddress.ip_interface(f"{addr_s}/32")
        desired_str = str(desired)
        host_str = str(desired.ip)

        # local dedupe (most important)
        if desired_str in seen_desired:
            continue
        seen_desired.add(desired_str)

        # cache hits
        if desired_str in IP_CACHE_BY_ADDRESS or host_str in IP_CACHE_BY_HOST:
            continue

        # exact match (use filter; get() can raise if >1)
        found = pick_best_ip(nb.ipam.ip_addresses.filter(address=desired_str), desired)
        if found:
            IP_CACHE_BY_ADDRESS[desired_str] = found
            IP_CACHE_BY_HOST[host_str] = found
            continue

        # host match
        candidates = list(nb.ipam.ip_addresses.filter(q=host_str))
        matched = None
        for c in candidates:
            try:
                ex = ipaddress.ip_interface(c.address)
            except Exception:
                continue
            if ex.ip == desired.ip:
                matched = c
                break

        if matched:
            # cache what we found; DO NOT auto-update mask here (can trigger duplicates)
            try:
                ex = ipaddress.ip_interface(matched.address)
                IP_CACHE_BY_ADDRESS[str(ex)] = matched
            except Exception:
                pass
            IP_CACHE_BY_HOST[host_str] = matched
            continue

        # Reserve in cache immediately so the same desired_str in this call won't be re-added
        IP_CACHE_BY_ADDRESS[desired_str] = "__PENDING__"
        IP_CACHE_BY_HOST[host_str] = "__PENDING__"

        needed.append({"address": desired_str, "status": "active"})

    if not needed:
        return

    if not commit:
        print(f"[DRY] Would bulk-create {len(needed)} IPs")
        return

    # Bulk create: NetBox supports list-of-dicts; PyNetBox returns list or single Record. :contentReference[oaicite:1]{index=1}
    created = nb.ipam.ip_addresses.create(needed)

    # Normalize return to iterable
    created_list = created if isinstance(created, list) else [created]

    for obj in created_list:
        # Some pynetbox versions return dict-like, others Record-like
        addr_val = obj.get("address") if isinstance(obj, dict) else getattr(obj, "address", None)
        if not addr_val:
            continue
        try:
            iface = ipaddress.ip_interface(str(addr_val).strip())
        except Exception:
            continue
        IP_CACHE_BY_ADDRESS[str(iface)] = obj
        IP_CACHE_BY_HOST[str(iface.ip)] = obj


# -----------------------------
# XLSX loading
# -----------------------------
def load_main_devices(ws_main) -> List[Dict[str, Any]]:
    devices = []
    for r in range(8, ws_main.max_row + 1):
        host_ip = ws_main.cell(row=r, column=1).value  # A
        if host_ip is None or str(host_ip).strip() == "":
            continue
        status_raw = ws_main.cell(row=r, column=2).value  # B
        if status_raw is None or str(status_raw).strip() == "":
            continue
        platform = ws_main.cell(row=r, column=3).value  # C
        name = ws_main.cell(row=r, column=10).value     # J
        model = ws_main.cell(row=r, column=11).value    # K

        if name is None or str(name).strip() == "":
            continue
        devices.append(
            {
                "host_ip": str(host_ip).strip(),
                "status_raw": "" if status_raw is None else str(status_raw).strip(),
                "platform": "" if platform is None else str(platform).strip(),
                "name": str(name).strip(),
                "model": "" if model is None else str(model).strip(),
            }
        )
    return devices

def load_interfaces_orig(ws_intf) -> Dict[str, List[Dict[str, Any]]]:
    by_host: Dict[str, List[Dict[str, Any]]] = {}
    for r in range(2, ws_intf.max_row + 1):
        hostname = ws_intf.cell(row=r, column=1).value
        if hostname is None or str(hostname).strip() == "":
            continue
        if_name = ws_intf.cell(row=r, column=2).value
        short_if = ws_intf.cell(row=r, column=3).value
        if if_name is None or str(if_name).strip() == "":
            continue
        rec = {
            "hostname": str(hostname).strip(),
            "name": str(if_name).strip(),
            "short_if": "" if short_if is None else str(short_if).strip(),  # <-- NEW
            "description": "" if ws_intf.cell(row=r, column=4).value is None else str(ws_intf.cell(row=r, column=4).value).strip(),
            "type_human": ws_intf.cell(row=r, column=5).value,
            "link": ws_intf.cell(row=r, column=7).value,
            "mode_raw": ws_intf.cell(row=r, column=10).value,
            "access_vlan_raw": ws_intf.cell(row=r, column=11).value,
            "tagged_vlans_raw": ws_intf.cell(row=r, column=14).value,
            "ip_address": ws_intf.cell(row=r, column=16).value,
            "mtu": ws_intf.cell(row=r, column=17).value,
        }
        by_host.setdefault(rec["hostname"], []).append(rec)
    return by_host

def load_interfaces(ws_intf) -> Dict[str, List[Dict[str, Any]]]:
    by_host: Dict[str, List[Dict[str, Any]]] = {}
    for r in range(2, ws_intf.max_row + 1):
        hostname = ws_intf.cell(row=r, column=1).value
        if hostname is None or str(hostname).strip() == "":
            continue
        if_name = ws_intf.cell(row=r, column=2).value
        short_if = ws_intf.cell(row=r, column=3).value
        if if_name is None or str(if_name).strip() == "":
            continue

        rec = {
            "hostname": str(hostname).strip(),
            "name": str(if_name).strip(),
            "short_if": "" if short_if is None else str(short_if).strip(),
            "description": "" if ws_intf.cell(row=r, column=4).value is None else str(ws_intf.cell(row=r, column=4).value).strip(),
            "type_human": ws_intf.cell(row=r, column=5).value,
            "link": ws_intf.cell(row=r, column=7).value,
            "mode_raw": ws_intf.cell(row=r, column=10).value,
            "access_vlan_raw": ws_intf.cell(row=r, column=11).value,
            "tagged_vlans_raw": ws_intf.cell(row=r, column=14).value,
            "mac": ws_intf.cell(row=r, column=15).value,      # ✅ MAC Add (O)
            "ip_address": ws_intf.cell(row=r, column=16).value,
            "mtu": ws_intf.cell(row=r, column=17).value,
        }
        by_host.setdefault(rec["hostname"], []).append(rec)
    return by_host

def load_inventory_type_overrides(ws_inv) -> Dict[Tuple[str, str], str]:
    """Return {(hostname, interface_name): description} from the Inventory sheet."""
    hdr = sheet_headers(ws_inv)

    def _col(*names: str) -> Optional[int]:
        for n in names:
            if n in hdr:
                return hdr[n]
        return None

    c_host = _col("Hostname", "Host", "Device Hostname")
    c_dev = _col("Device", "Interface", "Port")
    c_desc = _col("Description", "Type", "Media")

    if not (c_host and c_dev and c_desc):
        raise RuntimeError(
            f"Inventory sheet '{ws_inv.title}' must have headers for Hostname/Device/Description (found: {list(hdr)})"
        )

    out: Dict[Tuple[str, str], str] = {}
    for r in range(2, ws_inv.max_row + 1):
        host = ws_inv.cell(row=r, column=c_host).value
        dev = ws_inv.cell(row=r, column=c_dev).value
        desc = ws_inv.cell(row=r, column=c_desc).value
        if not host or not dev or not desc:
            continue
        out[(str(host).strip(), str(dev).strip())] = str(desc).strip()
    return out

def load_neighbors(ws, force_platform_blank: bool) -> Dict[str, List[Dict[str, str]]]:
    hdr = sheet_headers(ws)
    required = ["Hostname", "Remote Host", "MGMT IP", "Software"]
    if not force_platform_blank:
        required.append("Platform")
    for k in required:
        if k not in hdr:
            raise RuntimeError(f"Sheet '{ws.title}' missing required header '{k}'")

    out: Dict[str, List[Dict[str, str]]] = {}
    for r in range(2, ws.max_row + 1):
        host = ws.cell(row=r, column=hdr["Hostname"]).value
        if host is None or str(host).strip() == "":
            continue
        remote = ws.cell(row=r, column=hdr["Remote Host"]).value
        mgmt_ip = ws.cell(row=r, column=hdr["MGMT IP"]).value
        software = ws.cell(row=r, column=hdr["Software"]).value
        platform_val = ""
        if not force_platform_blank and "Platform" in hdr:
            platform = ws.cell(row=r, column=hdr["Platform"]).value
            platform_val = "" if platform is None else str(platform).strip()
        n = {
            "remote_host": "" if remote is None else str(remote).strip(),
            "mgmt_ip": "" if mgmt_ip is None else str(mgmt_ip).strip(),
            "platform": "" if force_platform_blank else platform_val,
            "software": "" if software is None else str(software).strip(),
        }
        out.setdefault(str(host).strip(), []).append(n)
    return out


# -----------------------------
# NetBox upsert logic
# -----------------------------
def upsert_device(nb, site, device_row: Dict[str, Any], commit: bool) -> Any:
    name = device_row["name"]
    model = device_row["model"] or "Unknown"
    platform = device_row["platform"]
    status = map_device_status(device_row.get("status_raw"))

    manufacturer_name = DEFAULT_MANUFACTURER_CISCO if is_cisco_platform(platform) else "Unknown"
    manufacturer = nb_get_or_create_manufacturer(nb, manufacturer_name, commit)
    device_type = nb_get_or_create_device_type(nb, manufacturer.id, model, commit)
    role_obj = nb_get_or_create_role(nb, DEFAULT_DEVICE_ROLE_NAME, commit)

    payload = {
        "name": name,
        "device_type": device_type.id,
        "role": role_obj.id,  # for new devices
        "site": site.id,
        "status": status,
        "custom_fields": {},
    }

    existing = nb.dcim.devices.get(name=name)
    if existing:
        if not commit:
            # Show role behavior
            cur_site = getattr(getattr(existing, "site", None), "name", None)
            if not getattr(existing, "role", None):
                print(f"[DRY] Would update device '{name}' (role unset; set -> {DEFAULT_DEVICE_ROLE_NAME})")
            else:
                print(f"[DRY] Would update device '{name}' (role set; leaving as-is)")
            return existing

        changed = False
        if existing.device_type and getattr(existing.device_type, "id", None) != device_type.id:
            existing.device_type = device_type.id
            changed = True
        #if getattr(existing, "site", None) and getattr(existing.site, "id", None) != site.id:
            #existing.site = site.id
            #changed = True
        if getattr(existing, "status", None) != status:
            existing.status = status
            changed = True

        # Set role only if currently unset
        if not getattr(existing, "role", None):
            existing.role = role_obj.id
            changed = True

        if changed:
            existing.save()
        return existing

    if not commit:
        print(f"[DRY] Would create device '{name}': {payload}")
        return type("Obj", (), {"id": -1, "name": name, "site": site, "custom_fields": {}})

    return nb.dcim.devices.create(**payload)

def prefetch_interfaces_for_device(nb, device_id: int) -> Dict[str, Any]:
    return {i.name: i for i in nb.dcim.interfaces.filter(device_id=device_id)}

def bulk_create_missing_interfaces(nb, device, intfs: List[Dict[str, Any]], existing_by_name: Dict[str, Any], inv_type_by_host_intf: Dict[Tuple[str, str], str], commit: bool) -> Dict[str, Any]:
    to_create: List[Dict[str, Any]] = []
    for row in intfs:
        name = row["name"]
        if name in existing_by_name:
            continue

        enabled = link_to_enabled(row.get("link"))
        inv_desc = inv_type_by_host_intf.get((row.get("hostname",""), name))
        nb_type = map_interface_type(row.get("type_human"), inventory_desc=inv_desc)
        mode = map_mode(row.get("mode_raw"))

        mtu = None
        mtu_val = row.get("mtu")
        if mtu_val is not None and str(mtu_val).strip() != "":
            try:
                mtu = int(mtu_val)
            except ValueError:
                mtu = None

        payload = {
            "device": device.id,
            "name": name,
            "type": nb_type,
            "enabled": enabled,
            "description": row.get("description") or "",
        }
        if mtu is not None:
            payload["mtu"] = mtu
        if mode is not None:
            payload["mode"] = mode

        to_create.append(payload)

    if not to_create:
        return existing_by_name

    if not commit:
        print(f"[DRY] Would bulk-create {len(to_create)} interfaces on {device.name}")
        return existing_by_name

    # Bulk create: one HTTP call
    nb.dcim.interfaces.create(to_create)
    # Re-fetch once to get ids
    return prefetch_interfaces_for_device(nb, device.id)

def upsert_interface_fields(nb_intf, desired: Dict[str, Any], commit: bool) -> None:
    if not commit:
        return
    changed = False
    for field in ("type", "enabled", "description", "mtu", "mode"):
        if field not in desired:
            continue
        new_val = desired[field]
        cur_val = getattr(nb_intf, field, None)
        if cur_val != new_val:
            setattr(nb_intf, field, new_val)
            changed = True
    if changed:
        nb_intf.save()

def set_interface_vlans_orig(nb, site, nb_intf, mode: Optional[str], access_vlan_raw: Any, tagged_vlans_raw: Any, commit: bool) -> None:
    if mode not in ("access", "tagged"):
        return

    desired_untagged: Optional[int] = None
    desired_tagged: List[int] = []

    if mode == "access":
        vid = parse_vlan_id(access_vlan_raw)
        if vid is not None:
            vlan = nb_get_or_create_vlan(nb, site.id, vid, commit)
            desired_untagged = vlan.id
        desired_tagged = []

    elif mode == "tagged":
        native_vid = parse_vlan_id(access_vlan_raw)
        if native_vid is not None:
            vlan = nb_get_or_create_vlan(nb, site.id, native_vid, commit)
            desired_untagged = vlan.id

        allowed = expand_vlan_list("" if tagged_vlans_raw is None else str(tagged_vlans_raw).strip())
        for vid in allowed:
            vlan = nb_get_or_create_vlan(nb, site.id, vid, commit)
            desired_tagged.append(vlan.id)

    if not commit:
        return

    # Compare current vs desired IDs
    cur_untagged = getattr(nb_intf, "untagged_vlan", None)
    cur_untagged_id = getattr(cur_untagged, "id", cur_untagged) if cur_untagged else None
    cur_tagged = getattr(nb_intf, "tagged_vlans", None) or []
    cur_tagged_ids = sorted([getattr(v, "id", v) for v in cur_tagged])
    desired_tagged_ids = sorted(desired_tagged)

    changed = False
    if cur_untagged_id != desired_untagged:
        nb_intf.untagged_vlan = desired_untagged
        changed = True
    if cur_tagged_ids != desired_tagged_ids:
        nb_intf.tagged_vlans = desired_tagged_ids
        changed = True
    if changed:
        nb_intf.save()

def set_interface_vlans(nb, site, nb_intf, mode: Optional[str], access_vlan_raw: Any, tagged_vlans_raw: Any, commit: bool) -> None:
    if mode not in ("access", "tagged"):
        return

    dev_name = getattr(getattr(nb_intf, "device", None), "name", "<unknown-device>")
    intf_name = getattr(nb_intf, "name", "<unknown-intf>")

    # Precompute some helpful context for logging
    native_str = "" if access_vlan_raw is None else str(access_vlan_raw).strip()
    tagged_str = "" if tagged_vlans_raw is None else str(tagged_vlans_raw).strip()

    try:
        desired_untagged: Optional[int] = None
        desired_tagged: List[int] = []

        if mode == "access":
            vid = parse_vlan_id(access_vlan_raw)
            if vid is not None:
                vlan = nb_get_or_create_vlan(nb, site.id, vid, commit)
                desired_untagged = vlan.id
            desired_tagged = []

        elif mode == "tagged":
            native_vid = parse_vlan_id(access_vlan_raw)
            if native_vid is not None:
                vlan = nb_get_or_create_vlan(nb, site.id, native_vid, commit)
                desired_untagged = vlan.id

            allowed = expand_vlan_list(tagged_str)
            for vid in allowed:
                vlan = nb_get_or_create_vlan(nb, site.id, vid, commit)
                desired_tagged.append(vlan.id)

    except (VlanScopeError, RuntimeError) as e:
        # Fail loudly for THIS interface, continue the rest of the run
        # - VlanScopeError: VID exists outside Global in strict mode
        # - RuntimeError: duplicates in Global (or any other explicit RuntimeError you raise)
        print(
            f"[SKIP] VLAN assignment for {dev_name}:{intf_name} "
            f"(mode={mode}, native/access={native_str!r}, tagged={tagged_str!r}) -> {e}"
        )
        return

    if not commit:
        return

    # Compare current vs desired IDs
    cur_untagged = getattr(nb_intf, "untagged_vlan", None)
    cur_untagged_id = getattr(cur_untagged, "id", cur_untagged) if cur_untagged else None
    cur_tagged = getattr(nb_intf, "tagged_vlans", None) or []
    cur_tagged_ids = sorted([getattr(v, "id", v) for v in cur_tagged])
    desired_tagged_ids = sorted(desired_tagged)

    changed = False
    if cur_untagged_id != desired_untagged:
        nb_intf.untagged_vlan = desired_untagged
        changed = True
    if cur_tagged_ids != desired_tagged_ids:
        nb_intf.tagged_vlans = desired_tagged_ids
        changed = True
    if changed:
        nb_intf.save()

def assign_ip_to_interface_orig(nb, nb_intf, ip_addr: str, commit: bool) -> None:
    """
    Assign an IP to an interface.

    Key fix:
      - Use nb.ipam.ip_addresses.update([...]) to PATCH only assignment fields,
        avoiding accidental address updates that can trigger duplicate errors.
    """
    if not ip_addr:
        return

    ip_obj = nb_get_or_create_ip(nb, ip_addr, commit)

    if not commit:
        return

    ip_id = getattr(ip_obj, "id", None)
    if not ip_id and isinstance(ip_obj, dict):
        ip_id = ip_obj.get("id")
    if not ip_id:
        raise RuntimeError(f"Cannot assign IP (no id): {ip_addr!r}")

    # Optional: fetch current to avoid unnecessary PATCH
    cur = nb.ipam.ip_addresses.get(ip_id)
    need = (
        getattr(cur, "assigned_object_type", None) != "dcim.interface"
        or getattr(cur, "assigned_object_id", None) != nb_intf.id
    )
    if not need:
        return

    nb.ipam.ip_addresses.update(
        [{
            "id": ip_id,
            "assigned_object_type": "dcim.interface",
            "assigned_object_id": nb_intf.id,
        }]
    )

def assign_ip_to_interface(nb, nb_intf, ip_addr: str, commit: bool) -> None:
    """
    Assign an IP to an interface.

    NetBox restriction:
      - You cannot reassign an IP address (change assigned_object) while that IP
        is designated as the device's primary_ip4.

    Fix:
      - If the IP is currently the device primary_ip4, temporarily clear it,
        patch the IP assignment, then allow later code (set_device_primary_ip4)
        to set it back.
    """
    if not ip_addr:
        return

    ip_obj = nb_get_or_create_ip(nb, ip_addr, commit)

    if not commit:
        return

    ip_id = getattr(ip_obj, "id", None)
    if not ip_id and isinstance(ip_obj, dict):
        ip_id = ip_obj.get("id")
    if not ip_id:
        raise RuntimeError(f"Cannot assign IP (no id): {ip_addr!r}")

    # Fetch current IP object to avoid unnecessary PATCH
    cur_ip = nb.ipam.ip_addresses.get(ip_id)
    need = (
        getattr(cur_ip, "assigned_object_type", None) != "dcim.interface"
        or getattr(cur_ip, "assigned_object_id", None) != nb_intf.id
    )
    if not need:
        return

    # Resolve device id from interface
    dev_ref = getattr(nb_intf, "device", None)
    dev_id = getattr(dev_ref, "id", dev_ref) if dev_ref else None

    # If some OTHER device is using this as primary_ip4, don't touch it
    # (avoids blowing up on shared/bad data)
    owners = list(nb.dcim.devices.filter(primary_ip4_id=ip_id))
    if owners and dev_id and owners[0].id != dev_id:
        print(
            f"[WARN] IP {ip_addr} is primary_ip4 for device '{owners[0].name}', "
            f"skipping assignment to {getattr(nb_intf, 'name', nb_intf.id)}"
        )
        return

    cleared_primary = False
    old_primary_id = None

    try:
        # If THIS device has this IP as primary, clear it temporarily
        if dev_id:
            dev = nb.dcim.devices.get(dev_id)
            prim = getattr(dev, "primary_ip4", None)
            prim_id = getattr(prim, "id", prim) if prim else None

            if prim_id == ip_id:
                old_primary_id = prim_id
                dev.primary_ip4 = None
                dev.save()
                cleared_primary = True

        # PATCH only the assignment fields
        nb.ipam.ip_addresses.update(
            [{
                "id": ip_id,
                "assigned_object_type": "dcim.interface",
                "assigned_object_id": nb_intf.id,
            }]
        )

    except Exception:
        # If we cleared primary and assignment failed, restore primary immediately
        if cleared_primary and dev_id and old_primary_id:
            try:
                dev = nb.dcim.devices.get(dev_id)
                dev.primary_ip4 = old_primary_id
                dev.save()
            except Exception:
                pass
        raise

def set_device_primary_ip4(nb, device, mgmt_ip_cidr: str, commit: bool) -> None:
    if not mgmt_ip_cidr:
        return
    ip_obj = nb_get_or_create_ip(nb, mgmt_ip_cidr, commit)
    if not commit:
        return
    dev = nb.dcim.devices.get(device.id)
    cur = getattr(dev, "primary_ip4", None)
    cur_id = getattr(cur, "id", cur) if cur else None
    if cur_id != ip_obj.id:
        dev.primary_ip4 = ip_obj.id
        dev.save()

def build_neighbors_json(device_name: str, mgmt_ip: str, cdp_by_host: Dict[str, List[Dict[str, str]]], lldp_by_host: Dict[str, List[Dict[str, str]]]) -> str:
    payload = {"device": device_name, "mgmt_ip": mgmt_ip, "cdp": cdp_by_host.get(device_name, []), "lldp": lldp_by_host.get(device_name, [])}
    return json.dumps(payload, indent=2, sort_keys=True)

def set_device_custom_field(nb, device_name: str, field: str, value: str, commit: bool) -> None:
    if not commit:
        return
    dev = nb.dcim.devices.get(name=device_name)
    if not getattr(dev, "custom_fields", None):
        dev.custom_fields = {}
    if dev.custom_fields.get(field) != value:
        dev.custom_fields[field] = value
        dev.save()

def get_current_time(str_option="dt"):
    """
    Captures the current time and returns it. Will return both date
    and time, or just one depending on the str_option provided.
    """
    now = datetime.now()
    str_option = str_option.lower()
    if str_option == "dt":
        return now.strftime("%m/%d/%Y") + ", " + now.strftime("%H:%M:%S")
    if str_option == "d":
        return now.strftime("%m/%d/%Y")
    if str_option == "t":
        return now.strftime("%H:%M:%S")
    return "Invalid selection.  Choose d for date, t for time, or dt for date + time."

def print_current_time():
    # Get the current time
    current_time = get_current_time("t")
    print("Current Time:", current_time)

def main() -> None:
    parser = argparse.ArgumentParser(description="Import devices/interfaces/IPs from XLSX into NetBox (optimized)")
    parser.add_argument("--xlsx", required=True, help="Path to XLSX file")
    parser.add_argument("--url", required=True, help="NetBox URL (e.g. http://netbox:8000)")
    parser.add_argument("--token", default=os.environ.get("NETBOX_TOKEN"), help="NetBox API token (or env NETBOX_TOKEN)")
    parser.add_argument("--site", default=DEFAULT_SITE_NAME, help=f"NetBox site name (default: {DEFAULT_SITE_NAME})")
    parser.add_argument("--dry-run", action="store_true", help="Print actions only; do not create/update in NetBox")
    parser.add_argument("--vlan-normalize", action="store_true", help="Normalize VLANs into Global: pick lowest id if duplicates in Global; \n move VLANs with matching VID from other groups/sites into Global (site=None) instead of creating a new VLAN."
)
    args = parser.parse_args()

    if not args.token:
        raise SystemExit("NetBox token not provided. Use --token or set NETBOX_TOKEN.")

    commit = not args.dry_run
    VLAN_BEHAVIOR = "normalize" if args.vlan_normalize else "strict"

    print_current_time()
    t0 = perf_counter()
    print(f"NetBox URL: {args.url}")
    print(f"XLSX: {args.xlsx}")
    print(f"Site: {args.site}")
    print("Mode:", "COMMIT" if commit else "DRY-RUN")

    # threading=True improves performance for large paginated fetches
    print("Setting netbox obj")
    nb = pynetbox.api(args.url, token=args.token, threading=True)

    print("Loading workbook obj")
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    print("Looping workbook obj")
    for required in ("Main", "Interfaces", "CDP", "LLDP"):
        if required not in wb.sheetnames:
            raise RuntimeError(f"Workbook missing sheet '{required}'")

    print("Checking workbook obj")
    ws_inv = wb["Inventory"] if "Inventory" in wb.sheetnames else None
    ws_main = wb["Main"]
    ws_intf = wb["Interfaces"]
    ws_cdp = wb["CDP"]
    ws_lldp = wb["LLDP"]

    print("Processing interfaces and links")
    interfaces_by_host = load_interfaces(ws_intf)
    cdp_links = build_cdp_links(ws_cdp)
    lldp_links = build_lldp_links(ws_lldp, interfaces_by_host)
    
    # Neighbor maps for interface descriptions
    cdp_port_remote = build_neighbors_from_cdp(ws_cdp)
    lldp_port_desc = build_neighbors_from_lldp(ws_lldp)
    
    # Neighbor maps for interface IP assignment (MGMT IP)
    cdp_port_mgmtip = build_port_mgmtip_from_cdp(ws_cdp)
    lldp_port_mgmtip = build_port_mgmtip_from_lldp(ws_lldp)

    print("Processing sites")
    site = nb_get_or_create_site(nb, args.site)

    devices = load_main_devices(ws_main)
    
    inv_type_by_host_intf = load_inventory_type_overrides(ws_inv) if ws_inv is not None else {}
    cdp_by_host = load_neighbors(ws_cdp, force_platform_blank=False)
    lldp_by_host = load_neighbors(ws_lldp, force_platform_blank=True)

    print(f"Found {len(devices)} device rows in Main (non-blank Main!A from row 8).")

    for d in devices:
        try:
            intfs = interfaces_by_host.get(d["name"], [])

            # Resolve mgmt IP to masked interface IP if possible; else fallback host/24.
            mgmt_ip, matched = resolve_mgmt_ip_with_mask(d["host_ip"], intfs)
            d["mgmt_ip"] = mgmt_ip

            # If no interface had the mgmt host IP, ensure we still assign the fallback IP
            # to some interface so primary_ip4 can be set.
            if not matched:
                target = prefer_mgmt_interface(intfs)
                if target is not None and (not target.get("ip_address")):
                    target["ip_address"] = mgmt_ip

            # Upsert device (primary_ip4 done later)
            dev = upsert_device(nb, site, d, commit)

            print(f"- {d['name']} ({d['mgmt_ip']}): {len(intfs)} interface rows")

            # If device is a dry-run stub, skip NetBox lookups requiring ids
            if not commit and getattr(dev, "id", None) in (-1, None):
                continue

            # Prefetch existing interfaces once per device
            existing_intfs_by_name = prefetch_interfaces_for_device(nb, dev.id)

            # Bulk-create missing interfaces
            existing_intfs_by_name = bulk_create_missing_interfaces(nb, dev, intfs, existing_intfs_by_name, inv_type_by_host_intf, commit)

            # Bulk ensure all IPs for this device exist (interface IPs + mgmt)
            ip_list: List[str] = []
            if d.get("mgmt_ip"):
                ip_list.append(d["mgmt_ip"])
            for row in intfs:
                if row.get("ip_address"):
                    ip_list.append(str(row["ip_address"]).strip())
            ip_list = list(dict.fromkeys(ip_list))  # preserves order, removes duplicates
            ensure_ips_exist_bulk(nb, ip_list, commit)

            # Update interface fields, VLANs, and IP assignments (only when changed)
            for row in intfs:
                name = row["name"]
                nb_intf = existing_intfs_by_name.get(name)
                if not nb_intf:
                    # Shouldn't happen after bulk create + prefetch, but safe
                    continue

                # --- Cable creation (CDP preferred, then LLDP) ---
                local_host = normalize_device_name(row.get("hostname", ""))
                local_full = row.get("name", "")
                local_short = row.get("short_if", "")

                remote = None

                # CDP keys use full local port names in your sheet (TenGigabitEthernet1/15)
                if local_host and local_full:
                    remote = cdp_links.get((local_host, local_full))

                # LLDP keys often use short local ports (Te1/15). Also supports MAC->port fallback.
                if not remote and local_host and local_short:
                    remote = lldp_links.get((local_host, local_short))

                if remote:
                    remote_host, remote_port = remote
                    remote_host = normalize_device_name(remote_host)

                    r_dev = nb.dcim.devices.get(name=remote_host)
                    if not r_dev:
                        # If NetBox stores the FQDN (rare), try original as fallback
                        r_dev = nb.dcim.devices.get(name=remote_host.strip())

                    if not r_dev:
                        print(f"[WARN] Remote device not found: {remote_host!r} (for {dev.name}:{nb_intf.name})")
                    else:
                        # Try remote port as-is, then expanded Cisco long form
                        cand_ports = [remote_port, expand_cisco_ifname(remote_port)]
                        r_intf = None
                        for p in cand_ports:
                            if not p:
                                continue
                            r_intf = nb.dcim.interfaces.get(device_id=r_dev.id, name=p)
                            if r_intf:
                                break

                        if not r_intf:
                            print(f"[WARN] Remote interface not found: {r_dev.name}:{remote_port!r} (for {dev.name}:{nb_intf.name})")
                        else:
                            ensure_cable_between_interfaces(nb, nb_intf, r_intf, commit)
                # --- end cable creation ---

                # --- Neighbor description overlay (LLDP first, then CDP) ---
                host = row.get("hostname", "")
                short_if = row.get("short_if", "")
                full_if = row.get("name", "")

                nbr_desc = None
                if host and short_if:
                    nbr_desc = lldp_port_desc.get(_key(host, short_if))
                if not nbr_desc and host and full_if:
                    nbr_desc = cdp_port_remote.get(_key(host, full_if))

                # Only overwrite if Interfaces.Description is blank (recommended)
                if nbr_desc and not (row.get("description") or ""):
                    row["description"] = nbr_desc
                # ---------------------------------------------------------------

                enabled = link_to_enabled(row.get("link"))
                inv_desc = inv_type_by_host_intf.get((row.get("hostname",""), name))
                nb_type = map_interface_type(row.get("type_human"), inventory_desc=inv_desc)
                mode = map_mode(row.get("mode_raw"))

                mtu = None
                mtu_val = row.get("mtu")
                if mtu_val is not None and str(mtu_val).strip() != "":
                    try:
                        mtu = int(mtu_val)
                    except ValueError:
                        mtu = None

                desired_fields: Dict[str, Any] = {
                    "type": nb_type,
                    "enabled": enabled,
                    "description": row.get("description") or "",
                }
                if mtu is not None:
                    desired_fields["mtu"] = mtu
                if mode is not None:
                    desired_fields["mode"] = mode

                upsert_interface_fields(nb_intf, desired_fields, commit)
                set_interface_vlans(nb, site, nb_intf, mode, row.get("access_vlan_raw"), row.get("tagged_vlans_raw"), commit)

                ip_val = row.get("ip_address")

                # If Interfaces sheet IP is blank, fall back to LLDP/CDP MGMT IP (if present)
                if not ip_val:
                    host = row.get("hostname", "")
                    short_if = row.get("short_if", "")
                    full_if = row.get("name", "")

                    nbr_ip = None
                    # LLDP uses Short IF
                    if host and short_if:
                        nbr_ip = lldp_port_mgmtip.get(_key(host, short_if))
                    # CDP uses full interface name
                    if not nbr_ip and host and full_if:
                        nbr_ip = cdp_port_mgmtip.get(_key(host, full_if))

                    if nbr_ip:
                        # MGMT IPs in CDP/LLDP sheets are typically host-only (no mask)
                        # Use /32 by default to avoid guessing the wrong subnet.
                        ip_val = f"{nbr_ip}/32" if "/" not in nbr_ip else nbr_ip

                if ip_val:
                    assign_ip_to_interface(nb, nb_intf, str(ip_val).strip(), commit)

            # Set primary_ip4 after assignments
            set_device_primary_ip4(nb, dev, d["mgmt_ip"], commit)

            # CDP/LLDP JSON custom field
            neighbors_json = build_neighbors_json(d["name"], d["mgmt_ip"], cdp_by_host, lldp_by_host)
            if not commit:
                print(f"[DRY] Would set custom_fields['cdp_neighbors_json'] on device {d['name']} (len={len(neighbors_json)} chars)")
            else:
                set_device_custom_field(nb, d["name"], "cdp_neighbors_json", neighbors_json, commit)
        except RequestError as e:
            # PyNetBox includes the response body with details like {'tagged_vlans': [...]}
            print(f"[SKIP] NetBox RequestError for device '{d.get('name', '<unknown>')}'. Skipping. Error: {e}")
            continue
        except Exception as e:
            # Optional: keep the run going even for unexpected errors
            print(f"[SKIP] Unexpected error for device '{d.get('name', '<unknown>')}'. Skipping. Error: {e}")
            continue
    delta_timer = perf_counter() - t0
    print("Done.")
    print(f"{delta_timer:.3f}s to finish")

if __name__ == "__main__":
    main()
