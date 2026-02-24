import os
import sys
import re
#import json
import orjson
from datetime import datetime
from pathlib import Path
from time import perf_counter

import openpyxl

VERBOSE = False

def load_device_type_cache(cache_path="device_type_cache.json"):
    print("Loading ", cache_path)
    try:
        #with open(cache_path, "r") as f:
            #return json.load(f)
        with open(cache_path, "rb") as f:
            return orjson.loads(f.read())
    except Exception:
        return {}

def save_device_type_cache(cache, cache_path="device_type_cache.json"):
    if VERBOSE:
        print("Saving ", cache_path)

    try:
        #with open(cache_path, "w") as f:
            #json.dump(cache, f, indent=2, sort_keys=True)
        with open(cache_path, "wb") as f:
            f.write(orjson.dumps(cache, option=orjson.OPT_INDENT_2))
    except Exception:
        pass

def find_val_in_col(value, column):
    """
    Returns the Row value in which the "value" was found
    """
    for n, cell in enumerate(column):
        if cell.value == value:
            return n + 1
    return 0

def gen_spacer(spacer_char="-", nl=2):
    """
    Returns a spacer string with 60 of designated character, "-" is default
    It will generate two lines of 60 characters
    """
    spacer = ""
    for i in range(nl):
        spacer += spacer_char * 60
        spacer += "\n"
    return spacer

def map_headers(wb_obj):
    """
    Returns a nested dictionary with the location and name of each
    column header
    """
    sheets = wb_obj.sheetnames
    return_value = {}
    ignore_this = ["Main", "Commands", "Settings", "Errors"]
    for sheet in sheets:
        if sheet not in ignore_this:
            row = wb_obj[sheet][1]
            sheet_mapper = {}
            for count, cell in enumerate(row):
                sheet_mapper[cell.value] = count + 1
            return_value[sheet] = sheet_mapper
    return return_value

def next_available_row(sheet_obj, col='A'):
    """
    Returns the number of the next available Row, it determines
    avaibaility based on whether there is an entry for hostname
    """
    column = sheet_obj[col]
    for index, cell in enumerate(column):
        if cell.value is None:
            return index + 1
    return len(column) + 1

def format_uptime(uptime):
    """
    Returns the Formated uptime.
    """
    str_years, str_weeks, str_days, str_hours, str_minutes = 0, 0, 0, 0, 0
    str_input = uptime.split(",")
    for i in str_input:
        i = i.strip()
        str_split = i.split(" ")
        if left(str_split[1], 3) == "yea":
            str_years = int(str_split[0])
        if left(str_split[1], 3) == "wee":
            str_weeks = int(str_split[0])
        if left(str_split[1], 3) == "day":
            str_days = int(str_split[0])
        if left(str_split[1], 3) == "hou":
            str_hours = int(str_split[0])
        if left(str_split[1], 3) == "min":
            str_minutes = int(str_split[0])
    if str_days > 365:
        years = str_days / 365
        if not years.is_integer():
            years = int(str(years).split(".")[0])
        str_days = str_days - years * 365
        str_years = str_years + years
    if str_days > 7:
        weeks = str_days / 7
        if not weeks.is_integer():
            weeks = int(str(weeks).split(".")[0])
        str_days = str_days - weeks * 7
        str_weeks = str_weeks + weeks
    if str_weeks > 52:
        years = str_weeks / 52
        if not years.is_integer():
            years = years.split(".")
            years = years[0]
        str_weeks = str_weeks - years * 52
        str_years = str_years + years

    return (str(str_years) + "y " +
            str(str_weeks) + "w " +
            str(str_days) + "d " +
            str(str_hours) + "h " +
            str(str_minutes) + "m "
            )

def left(s, amount):
    """
    Returns the left characters of amount size
    """
    return s[:amount]

def right(s, amount):
    """
    Returns the right characters of amount size
    """
    return s[-amount:]

def mid(s, offset, amount):
    """
    Returns the middle characters starting at offset of length amount
    """
    return s[offset:offset + amount]

def center_string(input_str, line_length=60):
    """
    Adds space in front of the string to center it.
    """
    rtr_str = ""
    extra_space = int((line_length - len(input_str)) / 2)
    if extra_space > 0:
        rtr_str = " " * extra_space + input_str
    return rtr_str

def rw_cell(sheet_obj, row, column, write=False, value=""):
    """
    Either writes or reads to/from a cell.
    """
    if write:
        sheet_obj.cell(row=row, column=column).value = value
        return None
    return sheet_obj.cell(row=row, column=column).value

def add_xls_tag(file_name):
    """
    Check the file_name to ensure it has ".xlsx" extension, if not add it
    """
    if file_name[:-5] != ".xlsx":
        return file_name + ".xlsx"
    else:
        return file_name

def get_xls_sheet(wb_obj, sheet_name):
    """
    Returns the Worksheet of provided Name
    """
    sheet_obj = wb_obj[sheet_name]
    sheet_obj.protection.sheet = False
    return sheet_obj

def open_xls(xls_input_file_name):
    """
    Returns the WorkBook of specified Name
    """
    try:
        return openpyxl.load_workbook(xls_input_file_name, data_only=True)
    except Exception as e:
        print(e)
        print("Please ensure the file exists or the correct filename was entered when utilizing the \"-i\" argument.")

def save_xls(wb_obj, input_file_name, file_name=None, output_dir=None, verbosity_level=0):
    """
    Saves the WorkBook to provided Directory and File Name.
    If no File Name and/or Directory provided it will save based on input_file_name.
    """
    file_save_string = ""
    if output_dir:
        output_dir = verify_path(output_dir)
        file_save_string = output_dir
    if file_name:
        file_save_string += file_name
    else:
        file_save_string += input_file_name[:-5] + "_new.xlsx"

    #current_time = get_current_time("t")
    #print(current_time, " - Saving the file to: " + file_save_string)
    t0 = perf_counter()
    wb_obj.save(file_save_string)
    delta_timer = perf_counter() - t0
    if verbosity_level >= 2:
        print(f"{delta_timer:.3f}s to save file: {file_save_string}")

def mod_dir_based_on_os(dir_name):
    """
    Based on type of system, it will change the '\\' to a '/', vice versa.
    if nt, then it assumes it is a windows.
    """
    if os.name == "nt":
        return dir_name.replace('/', "\\")
    return dir_name.replace('\\', "/")

def verify_path(output_dir):
    """
    Generates Path if it doesn't exist.
    """
    sys_type = os.name
    if sys_type == "nt" and ':' not in output_dir:
        output_dir = os.getcwd() + output_dir
    if not output_dir:
        output_dir = os.getcwd()
    if output_dir[-1] not in ["\\", "/"]:
        output_dir += "/"
    output_dir = mod_dir_based_on_os(output_dir)
    try:
        Path(output_dir).mkdir(parents=True, exist_ok=True)
    except Exception as e:
        if "Read-only" in str(e):
            print("Issue with path: " + output_dir)
            print("Error was raised: " + str(e))
            print("Suggestion: Try utilizing the full path.")
            if sys_type != "nt":
                print("If trying to use relative path please ensure that '/' is removed from leading directory name")
            print("Exiting now, please try again.")
            sys.exit()
        else:
            print("ERROR | ", e)
            sys.exit()
    return output_dir

def print_net_dev_msg(net_dev, msg):
    line1 = str(net_dev.main_col)
    if len(line1)<8:
        line1 += (8-len(line1))*" "
    line1 = " "+line1
    line2 = str(net_dev.host)
    if len(line2)<15:
        line2 += (15-len(line2))*" "
    print(line1, "|", line2, "|", msg)

def get_current_time(str_option="dt"):
    """
    Captures the current time and returns it. Will return both date
    and time, or just one depending on the str_option provided.
    """
    now = datetime.now()
    str_option = str_option.lower()
    if str_option == "dt":
        return now.strftime("%m/%d/%Y") + ", " + now.strftime("%H:%M:%S")
    if str_option == "d":
        return now.strftime("%m/%d/%Y")
    if str_option == "t":
        return now.strftime("%H:%M:%S")
    return "Invalid selection.  Choose d for date, t for time, or dt for date + time."